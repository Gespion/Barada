import email.utils
import datetime
import openpyxl
import smtplib
import pprint
import time
import ssl

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pymongo import MongoClient
#from barada.barada import settings
from funct import connect_to_mongodb, is_url_duplicate

filename = 'tuloss_contacts_test.xlsx'
mongo_uri = 'mongodb://localhost:27017'
mongo_db = 'tj-barada-db'

# Based on the currated contacts/jobs list, send batch emails per user
def sendjobs_to_users(curated_list):
    # Email parameters
    smtp_server = "YOUR-SMTP-SERVER-ADDRESS-HERE" #smtp.server.ext
    port = "YOUR-SMTP-PORT-HERE" 
    password = 'YOUR-SMTP-PASSWORD-HERE' 
    smtp_login = "YOUR-SMTP-LOGIN-HERE"
    sender_email = "YOUR-SENDER-EMAIL-HERE" #john.doe@email.ext
    sender_name = "YOUR-SENDER-NAME-HERE" #John Doe

    # Email content
    # Subject
    subject = 'Vos offres d\'emplois sur mesure'
    # Content header
    table_header = '<table style="background-color: #FFFFFF; padding: 5px; width: 80%"><tr style="color: white; background-color: #0496FF; text-align: left; font-family: sans-serif; font-size: 14px; width: 80%"><th style="border-radius: 2px; height: 18px; padding: 0px 10px; width: 35%">Soci&eacute;t&eacute;</th><th style="border-radius: 2px; height: 18px; padding: 0px 10px; width: 65%">Poste</th></tr>'
    email_header = '<html><head><title>Tuloss Jobs - Vos offres d\'emplois sur mesure</title></head><body><p style="color: #808080; font-family: sans-serif; font-size: 12px; padding: 5px 10px;">Hello, ci-dessous les offres correspondant &agrave; votre profil:</p>' + table_header
    email_footer = '</table><p style="color: #808080; font-family: sans-serif; font-size: 12px; padding: 5px 10px;">Cordialement,<br>L\'&eacute;quipe Tuloss Jobs<br><a href="https://jobs.tuloss.com">https://jobs.tuloss.com</a></p><p style="color: #808080; font-family: sans-serif; font-size: 10px; padding: 5px 10px; text-align: center;">Vous recevez cet email car vous &ecirc;tes abonn&eacute;(e) &agrave; notre alerte emplois.<br>Si vous avez des questions ou des remarques, n\'h&eacute;sitez pas à r&eacute;pondre &agrave; cet email.<br>Pour ne plus recevoir d\'emails de notre part, veuillez vous <a href="http://3v6p3.r.bh.d.sendibt3.com/mk/un/FRRNNhx8kaS1bw61r8CJddi9uWdnaD2bxB-S9dg567iW3TPFCHuAKA1SX_Eq-JJmVBHcY3V7US8DTl7RPX9L1krTNqbT80C0CRPL4_O7Sfh3vWPMSOZ4QSA8eLCvIU9JTNr8YV0">d&eacute;sinscrire</a>.</p></body></html>'
    message_html = email_header
    tr_td_open = '<tr style="color: #808080; font-family: sans-serif; font-size: 12px; border-radius: 3px;"><td style="padding: 0px 10px; border-bottom: 1px solid #0496FF;">'
    tr_td_close = '</td><td style="padding: 0px 10px; border-bottom: 1px solid #0496FF;">'
    a_link_open = '<a href="'
    a_link_close = '</a></td></tr>'
    # Plain text version to check later
    message_plain = 'Hello, ci-dessous les offres correspondant &agrave; votre profil:'

    # Preparing the server connection with SSL
    context = ssl.create_default_context()
    server = smtplib.SMTP_SSL(smtp_server, port, context=context)
    server.login(smtp_login, password)

    # Loop in contacts to send job offers
    i = 0
    #j = 0
    # pprint.pprint(curated_list)
    # Recover data from curated list
    for keys, details in curated_list.items():
        try:
            if i+1 in curated_list: # Check if index exist in dict to avoid Try/Except due to keyError and missing last item
                # If current and next contacts are the same (Trick to send batch jobs per user)
                if curated_list[i].get('contact_job_email') == curated_list[i+1].get('contact_job_email'):
                    # Create a job line
                    message_html = message_html + tr_td_open + details['contact_job_company'] + tr_td_close + a_link_open + details['contact_job_url'] + '">' + details['contact_job_title'] + a_link_close
                    i += 1  # Goto next job line
                else:
                    # If next user is different from current, add current line and close the job table
                    message_html = message_html + tr_td_open + details['contact_job_company'] + tr_td_close + a_link_open + details['contact_job_url'] + '">' + details['contact_job_title'] + a_link_close
                    message_html = message_html + email_footer

                    # Message header
                    msg = MIMEMultipart('alternative')
                    msg['From'] = email.utils.formataddr((sender_name, sender_email))
                    msg['Subject'] = subject

                    # Turn email content into plain/html MIMEText objects
                    part1 = MIMEText(message_plain, 'plain')
                    part2 = MIMEText(message_html, 'html')

                    # Add HTML/plain-text parts to MIMEMultipart message
                    # The email client will try to render the last part first
                    msg.attach(part1)
                    msg.attach(part2)

                    print('Sending to ' + details['contact_job_email'] + ' ...')
                    server.sendmail(sender_email, details['contact_job_email'], msg.as_string())
                    update_contacts_sent_date(filename,details['contact_job_email'],'Success')
                    #time.sleep(90)
                    print('Sent!')

                    # Reset the email header
                    message_html = email_header
                    i += 1
            else: # Here we're on the last dict item
                # If next user is different from current, add current line and close the job table
                message_html = message_html + tr_td_open + details['contact_job_company'] + tr_td_close + a_link_open + details['contact_job_url'] + '">' + details['contact_job_title'] + a_link_close
                message_html = message_html + email_footer

                # Message header
                msg = MIMEMultipart('alternative')
                msg['From'] = email.utils.formataddr((sender_name, sender_email))
                # msg['To'] = receiver_email
                msg['Subject'] = subject

                # Turn email content into plain/html MIMEText objects
                part1 = MIMEText(message_plain, 'plain')
                part2 = MIMEText(message_html, 'html')

                # Add HTML/plain-text parts to MIMEMultipart message
                # The email client will try to render the last part first
                msg.attach(part1)
                msg.attach(part2)

                print('Sending to ' + details['contact_job_email'] + ' ...')
                server.sendmail(sender_email, details['contact_job_email'], msg.as_string())
                update_contacts_sent_date(filename,details['contact_job_email'],'Success')
                #time.sleep(90)
                print('Sent!')
                # Reset the email header
                message_html = email_header
                i += 1
        except (KeyError, Exception) as ex:
            print('--- Oooops! Merci de vérifier ce qui cloche ---')
                      
            print("1.Exception: ",ex)
            print("2.Exception args: ",ex.args)
            #print("3.Message :",msg.as_string())
            print("4.Items :",msg.items())
            print("5.Params :",msg.get_params())

            print("User: ",details['contact_job_email'])
            update_contacts_sent_date(filename,details['contact_job_email'],'Fail')

            print('--- FIN DES DETAILS ---')
            break
    print("Congrats. All emails successfully sent!")
    server.close()

# Convert contact level to easy int
def contact_level_converter(level):
    switcher = {
        'inférieur au BAC': '0',
        'BAC': '1',
        'BAC + 2': '2',
        'BAC + 3': '3',
        'BAC + 4': '4',
        'BAC + 5 et plus': '5',
    }

    return switcher.get(level)

    #  Loop to get domain conversion from the dictionary switcher
    # args = args.split(",") # Transformation en liste
    # for arg in args:
    #    arg = arg.strip().lower()  # Remove useless spaces before and after the level
    # print("Arg: " + arg)
    #    clean_arg = switcher.get(arg, '0') # Call to switch level method, 0 is default
    #    if clean_arg not in level_list:  # Check to avoid duplicate in final level list
    #        level_list.append(clean_arg)  # If not already in the list then add
    # print("List item: " + str(level_list))

    # return str(level_list)


# Open the contact list file to collect user preference: level, domains and email
def read_contacts_list(contacts_file):
    # Read from Excel file
    wb = openpyxl.load_workbook(contacts_file)
    sheet = wb['tulossjobs']

    # Init the dictionary
    contacts = {}

    # Loop through values
    for row in range(2, sheet.max_row + 1):
        # print(row-2)
        # Init the contact dictionary line
        contacts[row - 2] = {}

        try:
            # Collect user email
            contacts[row-2]['user_email'] = sheet.cell(row=row, column=1).value
            # print(contacts[row-2]['user_email'])
            # Collect user level
            contacts[row-2]['user_level'] = contact_level_converter(sheet.cell(row=row, column=2).value)
            #print(contacts[row-2]['user_level'])
            domains = sheet.cell(row=row, column=3).value
            domains = domains.rstrip(";")
            domains = domains.strip()
            domains = list(domains.split(";"))
            for i in range(0, len(domains)):
                # Collect domain
                contacts[row-2]['user_domain_'+str(i)] = domains[i]
                # print(contacts[row-2]['user_domain_'+str(i+1)])

            # Collect user last jobs sent date
            contacts[row-2]['user_last_job_sent_date'] = sheet.cell(row=row, column=4).value
            #print(contacts[row-2]['user_last_job_sent_date'])
            #print(type(contacts[row-2]['user_last_job_sent_date']))
        except Exception as e:
            print("An exception occured when constructing user domains :",e)
            print("Faulty contact is :", contacts[row-2]['user_email'])
            print("Faulty domain is :", domains)
        else:
            pass

    return contacts

# Open the contact list file to update the last sent date
def update_contacts_sent_date(contacts_file,contact_email,status):
    # Read from Excel file
    wb = openpyxl.load_workbook(contacts_file)
    sheet = wb['tulossjobs']

    # Loop through values
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == contact_email and status == 'Success':
            sheet.cell(row=row, column=4).value = datetime.date.today()
            sheet.cell(row=row, column=5).value = ""
            wb.save(contacts_file)
        elif sheet.cell(row=row, column=1).value == contact_email and status == 'Fail':
            #print("User is: ", sheet.cell(row=row, column=1).value)
            #print("Current date is: ", sheet.cell(row=row, column=4).value)
            #print("Today's date is: ", datetime.date.today())
            sheet.cell(row=row, column=5).value = status
            wb.save(contacts_file)

# select user email and deconstruct the domains
def print_contacts(contacts):

    # for x in contacts:
    #   print(contacts[x])
    # for users and infos in dictionary
    for ide, user in contacts.items():
        print("User ID is : " + str(ide))
        # For keys and values in sub-dict info
        for key, val in user.items():
            print(key + " = " + val)

# Checking which company is not empty and retrieve it
#def is_company_null(company_b, company_a):
#    if company_b is None:
#        return company_a
#    return company_a

# Open the Job list file to collect Domains, Level and URL
def read_jobs_list(mongo_uri,jobs_db):
    global row_id

    # Create dict of sub-dicts
    # each sub-dict is based on domains collected from contact_list
    jobs = {}
    row_id = 0

    #data = connect_to_jobs_db(jobs_db)
    data = connect_to_mongodb(mongo_uri,jobs_db)

    # Loop through values to get jobs
    for item in data.find():
        # Init the dictionary
        jobs[row_id] = {}

        # Check if the job is still valid before matching and sending
        if item['job_expire_date'] >= datetime.datetime.today():
            # Site
            # job_site = item['job_site'].strip()
            #jobs[row_id]['job_site'] = item['job_site'] ##
            # print(jobs[row_id]['job_site'])

            # URL
            # job_url = item['job_url'].strip()
            jobs[row_id]['job_url'] = item['job_url']
            # print (jobs[row_id]['job_url'])

            # Date
            # job_date = item['job_date'].strip()
            jobs[row_id]['job_date'] = item['job_date']
            # print(jobs[row_id]['job_date'])

            # Expire Date
            # job_expire_date = item['job_expire_date'].strip()
            #jobs[row_id]['job_expire_date'] = item['job_expire_date'] ##
            # print(jobs[row_id]['job_expire_date'])

            # Title
            # job_title = item['job_title'].strip()
            jobs[row_id]['job_title'] = item['job_title']
            # print(jobs[row_id]['job_title'])

            # Company
            # job_company = item['job_company'].strip()
            jobs[row_id]['job_company'] = item['job_company'] #is_company_null(item['job_company_b'],item['job_company_a'])
            # print(jobs[row_id]['job_company'])

            domains = item['job_domain']
            sub_row_id = 0
            for domain in domains:
                # domain_item = domain.strip()
                jobs[row_id]['job_domain_' + str(sub_row_id)] = domain
                # print(str(sub_row_id) + ": " + jobs[row_id]['job_domain_' + str(sub_row_id)])
                sub_row_id += 1

            # Experience
            # job_experience = item['job_experience'].strip()
            #jobs[row_id]['job_experience'] = item['job_experience'] ##
            # print(jobs[row_id]['job_experience'])

            # Level
            # job_level = item['job_level'].strip()
            jobs[row_id]['job_level'] = item['job_level_num']
            # print(jobs[row_id]['job_level'])

            # Scrape Date
            # job_scrape_date = item['job_level'].strip()
            #jobs[row_id]['job_scrape_date'] = item['job_scrape_date'] ##
            #print(jobs[row_id]['job_scrape_date'])

            row_id += 1

    # print(jobs)

    return jobs


# Based on contacts/jobs files find matches
def match_contacts_jobs(contacts, jobs):

    # Init the variables
    matches = {}  # Final job list dict
    c = 0  # Final job list iterator

    # Loop through contacts
    for contacts_key, contacts_value in contacts.items():
        # print("---debut user---")
        a = 0  # contact domain iterator
        b = 0  # jobs domain iterator
        matches_temp = {}  # init/empty the temporary job list dict
        for c_k, c_v in contacts_value.items():
            if 'user_domain_'+str(a) in contacts_value and c_k == 'user_domain_'+str(a):
                # print(contacts_value['user_email'])
                # print(contacts_value['user_domain_'+str(a)])
                # print(contacts_value['user_level'])
                # print("raccourci: " + c_v)

                # Loop through jobs
                for jobs_key, jobs_value in jobs.items():
                    for j_k, j_v in jobs_value.items():
                        #  if job domain match contact domain and job scrape date is higher than the last job sent for the user
                        if (j_v == c_v) and (contacts_value['user_level'] in jobs_value['job_level']) and (jobs_value['job_date'] >= contacts_value['user_last_job_sent_date']):
                            # print("------debut jobs user---")
                            # print(jobs_value['job_level'])
                            # print(j_v)
                            # Check if job url isn't already stored in temp dict
                            if any(jobs_value['job_url'] in d.values() for d in matches_temp.values()):
                                # print('Doublon')
                                # print(jobs_value['job_url'])
                                pass
                            else:
                                # print('New')
                                # It's a new job offer, so we need to store in temp and final dict
                                matches_temp[b] = {}
                                matches_temp[b]['contact_job_email'] = contacts_value['user_email']
                                matches_temp[b]['contact_job_url'] = jobs_value['job_url']
                                matches_temp[b]['contact_job_title'] = jobs_value['job_title']
                                matches_temp[b]['contact_job_company'] = jobs_value['job_company']

                                matches[c] = {}
                                matches[c]['contact_job_email'] = matches_temp[b]['contact_job_email']
                                matches[c]['contact_job_url'] = matches_temp[b]['contact_job_url']
                                matches[c]['contact_job_title'] = matches_temp[b]['contact_job_title']
                                matches[c]['contact_job_company'] = matches_temp[b]['contact_job_company']

                                b += 1
                                c += 1
                            # print("------fin jobs user---")
                a += 1
        # print("---fin user---")
    # pprint.pprint(matches)
    return matches
###############################################################################################


# Executing the functions
# Read contacts list
contacts_list_dict = read_contacts_list(filename)
#print("--- Contacts ---")
#pprint.pprint(contacts_list_dict)
# Read jobs list
jobs_list_dict = read_jobs_list(mongo_uri,mongo_db)
#print("--- Jobs ---")
#pprint.pprint(jobs_list_dict)
# Match Contacts vs Jobs from previous files
contacts_jobs_list = match_contacts_jobs(contacts_list_dict, jobs_list_dict)
#print("--- Matches ---")
#pprint.pprint(contacts_jobs_list)
# Based on the matches list, send jobs to contacts
sendjobs_to_users(contacts_jobs_list)
